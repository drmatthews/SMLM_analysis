import os
from os import listdir
from os.path import isfile, join
import math
import numpy as np
import pandas as pd
from shapely.geometry import Polygon, MultiPoint, MultiLineString
from shapely.ops import cascaded_union, polygonize
from scipy.spatial import Delaunay
from sklearn.neighbors import KDTree
from sklearn.cluster import KMeans
from openpyxl import load_workbook
from read_roi import read_roi_file
from read_roi import read_roi_zip

from ..utils import (kdtree_nn,
                    polygon_area,
                    polygon_perimeter)
from ..utils import triangulation as tri

import warnings
warnings.filterwarnings("ignore")


class Cluster(object):
    def __init__(self, coords, cluster_id):
        N = coords.shape[0]
        self.x = np.reshape(coords[:, 0], (N, 1))
        self.y = np.reshape(coords[:, 1], (N, 1))
        self.cluster_id = cluster_id
        id_array = np.array(
            [cluster_id for i in range(coords.shape[0])],
            dtype=np.int32
        )
        self.id_array = np.reshape(id_array, (N, 1))
        self.occupancy = coords.shape[0]
        self.alpha = 0.01
        self.hull, edges = self._alpha_shape(coords, self.alpha)
        self.pc_ratio = None
        self.area = None
        self.perimeter = None
        self.nn = None
        self.center = None
        self.is_valid = False
        if edges:
            self.edges = np.vstack([np.reshape(a, (1, 4)) for a in edges])
            self.pc_ratio = self._pc(edges)
            self.area = self.hull.area
            self.perimeter = self.hull.length
            self.nn = np.reshape(self._near_neighbours(coords), (N, 1))
            self.center = self._center(np.hstack((self.x, self.y)))
            self.is_valid = True
        else:
            self.edges = edges

    def _alpha_shape(self, points, alpha):
        """
        Compute the alpha shape (concave hull) of a set
        of points.
        @param points: Numpy array of object coordinates.
        @param alpha: alpha value to influence the
            gooeyness of the border. Smaller numbers
            don't fall inward as much as larger numbers.
            Too large, and you lose everything!
        """
        if len(points) < 4:
            # When you have a triangle, there is no sense
            # in computing an alpha shape.
            return MultiPoint(list(points)).convex_hull

        def add_edge(edges, edge_points, coords, i, j):
            """
            Add a line between the i-th and j-th points,
            if not in the list already
            """
            if (i, j) in edges or (j, i) in edges:
                # already added
                return

            edges.add((i, j))
            edge_points.append(coords[[i, j]])

        # coords = numpy.array([point.coords[0] for point in points])
        coords = points
        tri = Delaunay(coords)
        edges = set()
        edge_points = []
        # loop over triangles:
        # ia, ib, ic = indices of corner points of the
        # triangle
        for ia, ib, ic in tri.vertices:
            pa = coords[ia]
            pb = coords[ib]
            pc = coords[ic]

            # Lengths of sides of triangle
            a = math.sqrt((pa[0] - pb[0])**2 + (pa[1] - pb[1])**2)
            b = math.sqrt((pb[0] - pc[0])**2 + (pb[1] - pc[1])**2)
            c = math.sqrt((pc[0] - pa[0])**2 + (pc[1] - pa[1])**2)
            # Semiperimeter of triangle
            s = (a + b + c) / 2.0
            # Area of triangle by Heron's formula
            area = math.sqrt(s * (s - a) * (s - b) * (s - c))

            if area > 0.0:
                circum_r = a * b * c / (4.0 * area)
                # Here's the radius filter.
                # print circum_r
                if circum_r < 1.0 / alpha:
                    add_edge(edges, edge_points, coords, ia, ib)
                    add_edge(edges, edge_points, coords, ib, ic)
                    add_edge(edges, edge_points, coords, ic, ia)
            else:
                continue
        m = MultiLineString(edge_points)
        triangles = list(polygonize(m))
        return cascaded_union(triangles), edge_points

    def _near_neighbours(self, points):
        tree = KDTree(points, leaf_size=2)
        dist, ind = tree.query(points[:], k=2)
        return dist[:, 1]

    def _pc(self, edges):
        e = np.vstack(edges)
        Y = np.c_[e[:, 0], e[:, 1]]
        pcr = tri.pc_ratio(Y)
        return pcr

    def _center(self, xy):
        kmeans = KMeans(n_clusters=1, random_state=0).fit(xy)
        return kmeans.cluster_centers_[0]


class ClusterList(object):
    def __init__(self):
        self.clusters = []
        self.n_clusters = len(self.clusters)
        self.noise = None

    def __getitem__(self, i):
        return self.clusters[i]

    def __len__(self):
        return self.n_clusters

    def append(self, item):
        self.clusters.append(item)
        self.n_clusters = len(self.clusters)

    def save(self, path, sheetname='image'):

        ext = os.path.splitext(path)[1]
        if ('xlsx' in ext):
            if self.clusters:

                writer = pd.ExcelWriter(path, engine='openpyxl')
                if os.path.exists(path):
                    book = load_workbook(path)
                    writer.book = book
                    writer.sheets = dict(
                        (ws.title, ws) for ws in book.worksheets
                    )

                coords = []
                stats = []
                hulls = []
                for c in self.clusters:
                    coords.append(np.hstack([c.x, c.y, c.nn, c.id_array]))
                    stats.append(np.hstack(
                        [c.area, c.perimeter, c.pc_ratio, c.occupancy]))
                    hull_cluster_id = np.array(
                        [c.cluster_id
                         for i in range(c.edges.shape[0])], dtype='int32')
                    hull_cluster_id = np.reshape(
                        hull_cluster_id, (c.edges.shape[0], 1))
                    hulls.append(np.hstack([c.edges, hull_cluster_id]))

                c_df = pd.DataFrame(np.vstack(coords))
                c_df.columns = [
                    'x [nm]', 'y [nm]', 'nn distance [nm]', 'cluster_id']
                s_df = pd.DataFrame(np.vstack(stats))
                s_df.columns = [
                    'area [nm^2]', 'perimeter [nm]', 'pc_ratio', 'occupancy']
                h_df = pd.DataFrame(np.vstack(hulls))
                h_df.columns = [
                    'x0 [nm]', 'y0 [nm]', 'x1 [nm]', 'y1 [nm]', 'cluster_id']

                c_df.to_excel(
                    writer,
                    sheet_name='{} cluster coordinates'.format(sheetname),
                    index=False
                )
                s_df.to_excel(
                    writer,
                    sheet_name='{} cluster statistics'.format(sheetname),
                    index=False
                )
                h_df.to_excel(
                    writer,
                    sheet_name='{} convex hulls'.format(sheetname),
                    index=False
                )

                if self.noise is not None:
                    noise_df = pd.DataFrame(self.noise)
                    noise_df.columns = ['x [nm]', 'y [nm]']
                    noise_df.to_excel(
                        writer,
                        sheet_name='{} noise'.format(sheetname),
                        index=False
                    )
                writer.save()
            else:
                print("no clusters to save")
        else:
            print("file path for saving must contain extension xlsx")


def collect_filenames(folder, dataset_name):
    fnames = []
    for filename in os.listdir(folder):
        if filename.endswith('xlsx') and (dataset_name in filename):
            fnames.append(filename)
    return fnames


def write_stats(output_path, data, dataset_name):
    if isinstance(data, dict):
        df = pd.DataFrame()
        for key, value in data.items():
            df[key] = value
    else:
        df = data

    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    if os.path.exists(output_path):
        book = load_workbook(output_path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, sheet_name=dataset_name, index=False)
    writer.save()


def collate_dataset_stats(out_dir, dataset_name, dataset_type, skip=[None]):
    print(dataset_name)
    df_list = []
    for filename in os.listdir(out_dir):
        if filename.endswith('xlsx') and (filename not in skip):
            if ((dataset_type in filename) and
                    (dataset_name in filename.lower())):

                df_list.append(
                    pd.read_excel(
                        os.path.join(out_dir, filename),
                        sheet_name='cluster statistics'
                    )
                )

    return pd.concat(df_list)


def cluster_near_neighbours(out_dir, dataset_name, dataset_type):
    print(dataset_name)
    nn_list = []
    for filename in os.listdir(out_dir):
        if filename.endswith('xlsx'):
            if (dataset_type in filename) and (dataset_name in filename):
                print(filename)
                df = pd.read_excel(
                    os.path.join(out_dir, filename),
                    sheet_name='cluster coordinates')

                nn = kdtree_nn(
                    df.as_matrix(columns=['x [nm]', 'y [nm]']))
                nn_list.append(nn)

    nn_df = pd.DataFrame(np.concatenate(nn_list, axis=0))
    nn_df.columns = ['Distance [nm]']

    return nn_df


def percentage_of_total(folder,
                        sheet_name,
                        dataset_name,
                        condition='eq',
                        cluster_column='lk'):
    percent = []
    for filename in os.listdir(folder):
        if filename.endswith('xlsx') and (dataset_name in filename):
            print(filename)
            df = pd.read_excel(os.path.join(folder, filename),
                               sheet_name=sheet_name)
            if 'eq' in condition:
                filtered = float(df[df[cluster_column] == -1].shape[0])
            elif 'gt' in condition:
                filtered = float(df[df[cluster_column] > -1].shape[0])
            print("number of filtered locs: {0}".format(filtered))
            total = float(df.shape[0])
            print("total number of localisations: {0}".format(total))
            p = filtered / total
            print("percentage: {0}".format(p))
            percent.append(p)
    return np.array(percent)


def percentage_objects_with_clusters(folder, sheet_name, dataset_name):
    percent = []
    for filename in os.listdir(folder):
        if filename.endswith('xlsx') and (dataset_name in filename):
            print(filename)
            df = pd.read_excel(os.path.join(folder, filename),
                               sheet_name=sheet_name)

            objects = float(df[df['object_id'] > -1].shape[0])
            clusters = float(df[df['cluster_id'] > -1].shape[0])

            print("number of object locs: {0}".format(objects))
            print("number of cluster locs: {0}".format(clusters))
            p = objects / clusters
            print("percentage: {0}".format(p))
            percent.append(p)
    return np.array(percent)


def mean_per_image(folder, dataset_name, parameter, sheet_name):
    mean = []
    for filename in os.listdir(folder):
        if filename.endswith('xlsx') and (dataset_name in filename):
            print(filename)
            df = pd.read_excel(
                os.path.join(folder, filename),
                sheet_name=sheet_name)
            if 'labels' in df:
                df = df[df['labels'] > -1]
            mean.append(df[parameter].mean())
    return np.array(mean)


def mean_near_neighbour_distance(folder, dataset_name, sheet_name):
    mean = []
    for filename in os.listdir(folder):
        if filename.endswith('xlsx') and (dataset_name in filename):
            print(filename)
            df = pd.read_excel(
                os.path.join(folder, filename),
                sheet_name=sheet_name)
            cluster_mean = []
            for cid, c in df.groupby('cluster_id'):
                cluster_mean.append(c['nn distance [nm]'].mean())
            mean.append(sum(cluster_mean) / len(cluster_mean))
    return np.array(mean)


def collect_stats(df, col):
    total = float(df.shape[0])
    not_noise = df[df[col] != -1]
    percentage = float(not_noise.shape[0]) / total
    objects_with_clusters = (
        float(df[(df['object_id'] != -1) & (df['cluster_id'] != -1)].shape[0])
    )
    labels = not_noise[col].unique()
    area = []
    perimeter = []
    occupancy = []
    stats = {}
    stats['label'] = labels
    for m in labels:
        group = df[df[col] == m]
        points = group.as_matrix(['x [nm]', 'y [nm]'])
        # area.append(group['area'].sum())
        area.append(polygon_area(points))
        perimeter.append(polygon_perimeter(points))
        occupancy.append(len(group.index))

    stats['area'] = area
    stats['perimeter'] = perimeter
    stats['occupancy'] = occupancy
    stats['percentage {0}s'.format(col[:len(col) - 3])] = percentage
    stats['not_noise'] = float(not_noise.shape[0])
    stats['objects_with_clusters'] = objects_with_clusters
    return stats


def import_cluster_stats(path, sheetname=None):
    if path.endswith('xlsx'):
        sn = 'image cluster stats'
        if sheetname:
            sn = '{0} cluster stats'.format(sheetname)
        try:
            df = pd.read_excel(path, sheet_name=sn)
            return df
        except ValueError:
            sn = 'image cluster stats'
            df = pd.read_excel(path, sheet_name=sn)
            return df
        except ValueError:
            return None
    else:
        raise ValueError("Input data must be in Excel format")


def batch_stats(folder, conditions, use_roi=False):

    for filename in os.listdir(folder):
        if filename.endswith('xlsx'):
            stats_path = os.path.join(folder, filename)
            basename = os.path.splitext(filename)[0]

            if use_roi:
                roi_zip_path = os.path.join(folder, basename +
                                            '_roiset.zip')
                roi_file_path = os.path.join(folder,
                                             basename + '_roiset.roi')
                if os.path.exists(roi_zip_path):
                    rois = read_roi_zip(roi_zip_path)
                elif os.path.exists(roi_file_path):
                    rois = read_roi_file(roi_file_path)
                else:
                    raise ValueError(("No ImageJ roi file exists -"
                                      "you should put the file in the same"
                                      "directory as the data"))

                stats = []
                for roi in rois.keys():
                    stats.append(import_cluster_stats(stats_path),
                                 sheetname=roi)
                stats_df = pd.concat(stats)
            else:
                stats_df = import_cluster_stats(stats_path)

            fnames = [f for f in listdir(folder) if isfile(join(folder, f))]
            outpath = os.path.join(folder, 'cluster_statistics_test.xlsx')
            for condition in conditions:
                condition_fnames = [fname for fname in fnames if condition in fname]
                cluster_stats = []
                for cf in condition_fnames:
                    cluster_stats.append(stats[cf])

                cddf = pd.DataFrame(condition_fnames)
                cddf.columns = ['filename']
                csdf = pd.concat(cluster_stats, axis=0)
                data = pd.concat([cddf, csdf.reset_index(drop=True)], axis=1)
                statistics.write_stats(stats_path, data, condition)
